import pandas as pd
from tkinter import Tk, Label, Text, Button, END, messagebox, Entry, Frame
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side


# Função para adicionar feedback e sugestão de melhoria
def add_feedback():
    global df  # Declarar df como global
    project_name = project_name_entry.get().strip()
    feedback = feedback_entry.get("1.0", END).strip()
    suggestion = suggestion_entry.get("1.0", END).strip()


    # Verificar se o nome do projeto foi fornecido
    if not project_name:
        messagebox.showwarning("Nome do Projeto", "Por favor, insira o nome do projeto.")
        return


    # Limitar o número de caracteres do feedback
    max_feedback_length = 500
    if len(feedback) > max_feedback_length:
        messagebox.showwarning("Feedback", f"O feedback deve ter no máximo {max_feedback_length} caracteres.")
        return


    if feedback:
        new_row = {
            'Nome do Projeto': project_name,
            'Data': datetime.now().strftime('%Y-%m-%d'),
            'Feedback': feedback,
            'Sugestão': suggestion
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)


        # Usar um nome de arquivo temporário para evitar problemas de permissão
        temp_filename = 'feedbacks_temp.xlsx'
        df.to_excel(temp_filename, index=False)


        # Aplicar formatação ao arquivo Excel
        format_excel(temp_filename, 'feedbacks.xlsx')


        # Remover o arquivo temporário após 2 segundos
        root.after(2000, lambda: os.remove(temp_filename) if os.path.exists(temp_filename) else None)


        messagebox.showinfo("Feedback", "Feedback e sugestão adicionados com sucesso!")
    else:
        messagebox.showwarning("Feedback", "Por favor, insira o feedback.")


def format_excel(input_filename, output_filename):
    # Carregar o arquivo temporário
    wb = load_workbook(input_filename)
    ws = wb.active


    # Definir estilos
    header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))


    # Aplicar estilo ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border


    # Aplicar bordas e definir largura das colunas
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


    # Ajustar largura das colunas
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                column_widths[cell.column] = max((column_widths.get(cell.column, 0), len(str(cell.value))))


    for col, width in column_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width + 2  # +2 para algum padding


    # Salvar o arquivo formatado
    wb.save(output_filename)


# Inicializar DataFrame
try:
    df = pd.read_excel('feedbacks.xlsx')
except FileNotFoundError:
    df = pd.DataFrame(columns=['Nome do Projeto', 'Data', 'Feedback', 'Sugestão'])


# Função para validar a entrada de texto
def validate_text_input(event):
    char = event.char
    if not (char.isalnum() or char.isspace() or char in ["", "\b", "\x7f"]):  # Permitir letras, números, espaços, Backspace e Delete
        return "break"  # Bloqueia a entrada do caractere


def focus_next_widget(event):
    # Mudar o foco para o próximo campo
    next_widget = root.focus_get().tk_focusNext()
    if next_widget:
        next_widget.focus()
    return "break"  # Impedir o comportamento padrão do Tab


# Função para limpar os campos de entrada
def clear_entries():
    project_name_entry.delete(0, END)
    feedback_entry.delete(1.0, END)
    suggestion_entry.delete(1.0, END)


root = Tk()
root.title("Coleta de Feedbacks")


# Configurar cor de fundo da janela
root.configure(bg='#493d57')  # Cor do fundo


# Configurar a grade
root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(1, weight=1)
root.grid_rowconfigure(2, weight=1)
root.grid_rowconfigure(3, weight=1)
root.grid_rowconfigure(4, weight=1)
root.grid_rowconfigure(5, weight=1)
root.grid_rowconfigure(6, weight=1)
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=1)


# Criar e posicionar widgets
Label(root, text="Digite o nome do projeto:", bg='#493d57', fg='#fcdb04').grid(row=0, column=0, columnspan=2, sticky='w', padx=10, pady=5)
project_name_entry = Entry(root)
project_name_entry.grid(row=1, column=0, columnspan=2, sticky='ew', padx=10, pady=5)
project_name_entry.bind('<Tab>', focus_next_widget)
project_name_entry.bind("<KeyPress>", validate_text_input)


Label(root, text="Digite seu feedback (máximo de 500 caracteres):", bg='#493d57', fg='#fcdb04').grid(row=2, column=0, columnspan=2, sticky='w', padx=10, pady=5)
feedback_entry = Text(root, height=10, width=50, wrap='word')
feedback_entry.grid(row=3, column=0, columnspan=2, sticky='ew', padx=10, pady=5)
feedback_entry.bind('<Tab>', focus_next_widget)
feedback_entry.bind("<KeyPress>", validate_text_input)


Label(root, text="Digite sua sugestão de melhoria (opcional):", bg='#493d57', fg='#fcdb04').grid(row=4, column=0, columnspan=2, sticky='w', padx=10, pady=5)
suggestion_entry = Text(root, height=5, width=50, wrap='word')
suggestion_entry.grid(row=5, column=0, columnspan=2, sticky='ew', padx=10, pady=5)
suggestion_entry.bind('<Tab>', focus_next_widget)
suggestion_entry.bind("<KeyPress>", validate_text_input)


# Frame para centralizar os botões
button_frame = Frame(root, bg='#493d57')
button_frame.grid(row=6, column=0, columnspan=2, pady=10)


# Botões
Button(button_frame, text="Enviar feedback", command=add_feedback, bg='#219cb6', fg='#fcdb04').grid(row=0, column=0, padx=10)
Button(button_frame, text="Limpar", command=clear_entries, bg='#219cb6', fg='#fcdb04').grid(row=0, column=1, padx=10)


# Definir o foco inicial no campo de Nome do Projeto
root.after(100, lambda: project_name_entry.focus())


root.mainloop()


